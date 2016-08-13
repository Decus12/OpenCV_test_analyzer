import numpy as np
#import imutils
import cv2
import os
import xlsxwriter
from openpyxl import load_workbook
import argparse
import time

timestamp1 = time.time()
 
# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-d", "--dir", required=True,
	help="path to the input directory")
ap.add_argument("-t", "--tl", required=True,
	help="T-line's distance from C-line in pixels")
ap.add_argument("-tem", "--template", required=True,
	help="Template for test area")

args = vars(ap.parse_args())
tl = int(args["tl"])
testArea = args["template"]
if args["dir"] is None:
    print "Please give an image folder as argument"
else:
    indir = args["dir"]
    #execfile ("multi_img.py")

"""***************************************************************
**************************Excel Reading***************************
***************************************************************"""
#Reading class numbers and their thresholds. Similar to reading in classifier_teach.py
wb = load_workbook(filename = 'ranges.xlsx')
sheet = wb['Ref Ranges']
Refs = np.array(sheet.columns[0])
Thres = np.array(sheet.columns[1])


refSize = np.shape(Refs)
classRefs = np.zeros(refSize[0])
classThres = np.zeros(refSize[0])

for c in range(1,refSize[0]):
    classRefs[c] = Refs[c].value
    
for c in range(1,refSize[0]):
    classThres[c] = Thres[c].value
    
classRefs = classRefs[1:refSize[0]-1] #Remove the column name cell's "value" from vector and last cell, since it's empty in this case   
classThres = classThres[1:refSize[0]+1] #Remove the column name cell's "value" and add one more cell
#to the end because this vector is larger due to needed additional value for the last class.    
print classRefs
print classThres

caller = 1 #Used for telling line analysis algorithm what excel file to write.
execfile ("multi_img.py")

#Reading the ratio results from the algorithm executed above.
wb = load_workbook(filename = 'Classified tests.xlsx')
sheet = wb['Test results']
cellsRatio = np.array(sheet.columns[7])
cellsSize = np.shape(cellsRatio)
ratios = np.zeros(cellsSize[0])

for r in range(1,cellsSize[0]):
    ratios[r] = cellsRatio[r].value
ratios = ratios[1:cellsSize[0]]

classes = np.zeros(cellsSize[0]-1)
#print ratios
#print cellsSize

"""***************************************************************
**************************Classifying*****************************
***************************************************************"""
for i in xrange(0,cellsSize[0]-1): #loop through all image ratios
    for j in xrange(0,refSize[0]-1): #loop through all class thresholds for each image
        if ratios[i] >= classThres[j] and ratios[i] < classThres[j+1]: #Determine which class' thresholds match the image's ratio.
            classes[i] = classRefs[j]
            break
            
#print classes

"""***************************************************************
**************************Excel Writing***************************
***************************************************************"""
workbook = xlsxwriter.Workbook('classified.xlsx')
worksheet = workbook.add_worksheet('classifier results')
worksheet.write('A1', 'Referance')
worksheet.write('B1', 'Ratio')
worksheet.set_column('A:B', 20)

n = 0
for i in xrange(0,cellsSize[0]-1):
    n = n+1
    
    worksheet.write(n, 0, classes[i])
    worksheet.write(n, 1, ratios[i])

workbook.close()

execfile ("accuracy.py")
    
    
timestamp2 = time.time()
print "time", (timestamp2 - timestamp1)