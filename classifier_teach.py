import numpy as np
#import imutils
import cv2
import os
import xlsxwriter
from openpyxl import load_workbook
import argparse
import time

timestamp1 = time.time()
 
# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()
ap.add_argument("-d", "--dir", required=True,
	help="path to the input directory")
ap.add_argument("-t", "--tl", required=True,
	help="T-line's distance from C-line in pixels")
ap.add_argument("-tem", "--template", required=True,
	help="Template for test area")

args = vars(ap.parse_args())
tl = int(args["tl"])
testArea = args["template"]
if args["dir"] is None:
    print "Please give an image folder as argument"
else:
    indir = args["dir"]
    caller = 0 #Used for telling line analysis algorithm what excel file to write.
    execfile ("multi_img.py")

    
"""***************************************************************
**************************Excel Reading***************************
***************************************************************"""
#Using openpyxl read the wanted coloumns from excel file and create empty vectors of their size.
wb = load_workbook(filename = 'tests.xlsx')
sheet = wb['Test results']
cellsRef = np.array(sheet.columns[1])
cellsRatio = np.array(sheet.columns[7])
cellsSize = np.shape(cellsRatio)
references = np.zeros(cellsSize[0])
ratios = np.zeros(cellsSize[0])

#Getting the actual values stored in excel cells with .value
for c in range(1,cellsSize[0]):
    references[c] = cellsRef[c].value
#print references
for r in range(1,cellsSize[0]):
    ratios[r] = cellsRatio[r].value
ratios = ratios[1:cellsSize[0]] #Remove the column name cell's "value" from vector
#print 'ratios', ratios

"""***************************************************************
*****************Number of each reference class*******************
***************************************************************"""
#Counting the number of each reference value in case there are uneven number of diffrent reference classes.
refmax = int(max(references)) # How many classes there are at default
refCount = np.zeros(refmax+1) #Empty vector for each reference's count
refNumber = 1

for s in range(0,refmax+1): 
    for n in range(1,cellsSize[0]):
        if references[n] == refNumber:
            #When reference equals reference number, increase it's count in matching point of count vector
            refCount[refNumber] = refCount[refNumber] + 1  
            
    refNumber = refNumber + 1
refCount = refCount[1:refmax+1]

#print 'refcount', refCount



"""***************************************************************
************Class average ratios and standard deviation***********
***************************************************************"""

#Based on the number of each reference class' count crop the approriate part of the ratios values
startPoint = 0
ratiosSd = np.zeros(refmax) #Empty vector for each class' standard deviation
ratiosAvg = np.zeros(refmax) #Empty vector for each class' average ratio
for i in range(0,refmax):
    ratioRefRange = ratios[int(startPoint):int(startPoint+refCount[i])] #Ratios variable values
    #changes due to some weird reason here. It works for this loop, but if used later, it must be
    #read from excel file again
    #print 'ratios3', ratios 
    ratioAvg = np.average(ratioRefRange)
    ratioSdRange = ratioRefRange
    
    #Calculate standard deviation from each classes average value.
    for j in range (0,np.shape(ratioRefRange)[0]):
        ratioSdRange[j] = (ratioRefRange[j]-ratioAvg)**2   
    ratioSd = np.sqrt(np.average(ratioSdRange))
    

    #print ratioSd, ratioAvg
    startPoint = startPoint + refCount[i] #Move crop point by the amount that was croped for this pass
    
    #Write standard deviation and average ratio for each class to their matching points.
    ratiosSd[i] = ratioSd
    ratiosAvg[i] = ratioAvg

#print ratiosSd
#print ratiosAvg

#Based on standard deviation and average ratio of each class determine whether two classes are too close to each other.
classNumber = 1 #Written to append mark vector if append is necessary
classThresholds = np.zeros(refmax) #append mark vector

grown = 0
for i in xrange(0,refmax-1):
    halfAvg = (ratiosAvg[i+1]-ratiosAvg[i])/2 #Mid point between average values
    #print 'halfAvg', halfAvg
    #print ratiosSd[i]
    if halfAvg < ratiosSd[i]: #If sd is larger than mid point between values the two classes need to be appended
        #Append marks are written to two points on both sides of the mid point. This requires if
        #structure so that appending isn't overwritten in case when loop passes to next point and
        #it would also result in appending with next class.
        #print classThresholds[i]
        if classThresholds[i] == 0: 
            classThresholds[i] = classNumber
            classThresholds[i+1] = classNumber
            grown = 0
            #print classThresholds
        else:
            classNumber = classNumber + 1 #Raise the class number by one, so that different
            #appending points can be separated from each other in case more than one appending must be done.
            grown = 1
    elif halfAvg > ratiosSd[i] and classThresholds[i] > 0 and grown == 0:
        classNumber = classNumber + 1
        grown = 0
    elif halfAvg == ratiosSd[i]:
        classThresholds[i] = classNumber
        classThresholds[i+1] = classNumber
        classNumber = classNumber + 1
        grown = 1
        
        
print classThresholds
"""***************************************************************
************************Class appending***************************
***************************************************************"""

#If appending points were found then we perform this operation. Otherwise we can just move on to excel writing part.

if max(classThresholds):
    ThresMax = int(max(classThresholds)) #How many appendings must be done

    ratiosAvg = np.zeros(refmax-ThresMax)
    ratiosSd = np.zeros(refmax-ThresMax)
    
    #Must load ratios again because of variable values changing for some reason earlier.
    ratios = np.zeros(cellsSize[0])

    for r in range(1,cellsSize[0]):
        ratios[r] = cellsRatio[r].value
    ratios = ratios[1:cellsSize[0]]
    #print ratios

    #Using the same calculations we recalculate average ratios and their sd's by croping larger
    #areas from ratios based on appending marks.
    startPoint = 0
    appender = 0 #Used to move writing points behind since number of classes is getter smaller
    for k in xrange(0,refmax):
        if classThresholds[k] == 0: #if there's no need for appending at class point then this is
            #very similar to earlier average and sdcalculation. Except that croping point is moved
            #behind by appender. It's 0 until loop reaches a point where it needs to append, so it
            #doesn't interfere until after some appending has happened. 
            ratioRefRange = ratios[int(startPoint):int(startPoint+refCount[k-appender])]
            #print ratioRefRange
            ratioAvg = np.average(ratioRefRange)
            ratioSdRange = ratioRefRange

            for j in range (0,np.shape(ratioRefRange)[0]):
                ratioSdRange[j] = (ratioRefRange[j]-ratioAvg)**2   
            ratioSd = np.sqrt(np.average(ratioSdRange))


            #print 'jee', ratioAvg
            startPoint = startPoint + refCount[k-appender]

            ratiosSd[k-appender] = ratioSd
            ratiosAvg[k-appender] = ratioAvg
        elif classThresholds[k] > 0 and classThresholds[k] > appender: #When appending needs to be
            #done this is done. Croping area is now the area of both original classes being appended.  
            ratioRefRange = ratios[int(startPoint):int(startPoint+refCount[k]+refCount[k+1])]
            #print ratioRefRange
            ratioAvg = np.average(ratioRefRange)
            ratioSdRange = ratioRefRange

            for j in range (0,np.shape(ratioRefRange)[0]):
                ratioSdRange[j] = (ratioRefRange[j]-ratioAvg)**2   
            ratioSd = np.sqrt(np.average(ratioSdRange))


            #print 'jee2', ratioAvg
            startPoint = startPoint + refCount[k]+refCount[k+1]
            


            ratiosSd[k-appender] = ratioSd
            ratiosAvg[k-appender] = ratioAvg
            appender = appender + 1 #Since some classes disappear but we need to loop as many times
            #as there are original classes we grow appender for each appending done so that writing
            #of the new values is done to the next empty point in new sd and average vectors.



    print ratiosSd
    print ratiosAvg        
        

else:
    ThresMax = 0 #If no appending was necessary then we don't need to make final threshold vector smaller.
"""***************************************************************
**************************Excel Writing***************************
***************************************************************"""
classThres = np.zeros(refmax-ThresMax)
lastAvg = 0 #Needed for calculating the first class' threshold correctly since it doesn't have a previous value to compare it to, so ratiosAvg[x-1] doesn't work for it.

#Calculate the threshold where each class begins
for x in xrange(0,refmax-ThresMax):
    classThres[x] = (ratiosAvg[x]-lastAvg )/2+lastAvg 
    lastAvg = ratiosAvg[x]
    
print classThres

workbook = xlsxwriter.Workbook('ranges.xlsx')
worksheet = workbook.add_worksheet('Ref Ranges')
worksheet.write('A1', 'Referance')
worksheet.write('B1', 'Class Threshold')
worksheet.write('C1', 'Class avg')
worksheet.write('D1', 'Class sd')
worksheet.write('E1', 'Appended')
worksheet.set_column('B:D', 20)

n = 0
for i in range(0,refmax-ThresMax):
    
    n = n+1
    
    worksheet.write(n, 0, n)
    worksheet.write(n, 1, classThres[i])
    worksheet.write(n+1, 1, 10.0) #Write additional very high ratio cap to last class for classifier
    worksheet.write(n, 2, ratiosAvg[i]) #Not needed for classifier but writen for reference
    worksheet.write(n, 3, ratiosSd[i]) #Same as average ratio
n = 0
for j in range(0,refmax):
    
    n = n+1    
    worksheet.write(n, 4, classThresholds[j])

workbook.close()

timestamp2 = time.time()
print "time", (timestamp2 - timestamp1)
        
