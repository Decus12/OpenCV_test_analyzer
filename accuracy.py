import numpy as np
import cv2
from openpyxl import load_workbook

"""***************************************************************
**************************Excel Reading***************************
***************************************************************"""
#Using openpyxl read the wanted coloumns from excel file and create empty vectors of their size.

#Original reference values
wb = load_workbook(filename = 'tests.xlsx')
sheet = wb['Test results']
cellsRef = np.array(sheet.columns[1])
cellsSize = np.shape(cellsRef)
references = np.zeros(cellsSize[0])

#Getting the actual values stored in excel cells with .value
for c in range(1,cellsSize[0]):
    references[c] = cellsRef[c].value
references = references[1:cellsSize[0]]    
#print references

#Classifier's class results
wb2 = load_workbook(filename = 'classified.xlsx')
sheet2 = wb2['classifier results']
cellsRef2 = np.array(sheet2.columns[0])
cellsSize2 = np.shape(cellsRef2)
references2 = np.zeros(cellsSize2[0])

for c in range(1,cellsSize2[0]):
    references2[c] = cellsRef2[c].value
references2 = references2[1:cellsSize2[0]]    
#print references2

#Class appending results
wb3 = load_workbook(filename = 'ranges.xlsx')
sheet3 = wb3['Ref Ranges']
cellsRef3 = np.array(sheet3.columns[4])
cellsSize3 = np.shape(cellsRef3)
references3 = np.zeros(cellsSize3[0])

for c in range(1,cellsSize3[0]):
    references3[c] = cellsRef3[c].value
#references3 = [ 0.,  1.,  1.,  0.,  0.,  2.,  2.,  0.,  0.,  0.,  0.] #For testing loop below
#print references3

"""***************************************************************
*************************Changing classes*************************
***************************************************************"""
if max(references3) == 0:
    pass
else:
    ref = 1 #Used to compare and write reference results during the loop
    refsubs = 0 #Used for substracting from ref when there is no appending done to a class but total class count is lower than original
    refadd = 0 #Used to add to ref when appending is done to a class but there are non-appended classes before it


    for x in xrange(1,cellsSize3[0]-1): #Go through all original classes

        for y in xrange(0,cellsSize[0]-1): #Go through all results 
            if references[y] == ref and references3[x] > 0: #If There is appending done rewrite class reference with ref + refadd
                # so it's value is correct even after non appending classes precede it.
                references[y] = references3[x] + refadd
            elif references[y] == ref and references3[x] == 0: #If Appending is not done then rewrite class with ref - refsubs
                # so that it's value is correct even after class appending has been done
                references[y] = ref-refsubs

        #Compare reference class to the next and if it's the same and class is appended then add 1 to refsubs. Else when appending is not doen add 1 to refadd
        if references3[x] == references3[x+1] and references3[x] > 0: 
            refsubs = refsubs + 1
        elif references3[x] == 0:
            refadd = refadd +1




        ref = ref + 1
        #Used for the last class. Because it can't be compared to next class it's value must be adjusted differently.     
        if ref == cellsSize3[0]-1 and references3[x] == 0:
            for z in xrange(0,cellsSize[0]-1):
                if references[z] == ref:        
                    references[z] = ref-refsubs
#print references

"""***************************************************************
***********************Accuracy calculation***********************
***************************************************************"""
refResults = abs(references - references2) #Substract classifier results from adjusted classes
#This way only non matching ones are left larger than 0
#print refResults
count = 0.0
for s in xrange (0,cellsSize[0]-1): #Must use for-loop, because sum + dividing results in NaN when 0 is divided by 0
    if refResults[s] > 0:
        count = count + 1
#print count

accuracy = 1-count/(cellsSize[0]-1)
print 'accuracy', accuracy*100,'%'
